# -*- coding: utf-8 -*-
"""
Created on Thu Jul 21 12:40:46 2022

@author: rtse
"""

import openpyxl as xl
import pandas as pd
from typing import List, Dict, Tuple
from abc import abstractmethod, ABC

from shared_code.ExcelFiller.FillerInput import FillerInputModel
    

class RateTabFiller(ABC):
    
    def __init__(self, rate_start_row: int,
                 rate_start_col: int, name_cells: list,
                 source_tab: str):
        self.rate_start_row = rate_start_row
        self.rate_start_col = rate_start_col
        self.name_cells = name_cells
        self.source_tab = source_tab
        
    def fill_tab(self, wb: xl.workbook.workbook.Workbook, 
                   rates: FillerInputModel, target_tab:str = None,
                   **kwargs) \
        -> xl.workbook.workbook.Workbook:
        if target_tab == None:
            rate_sheet = wb[self.source_tab]
        elif target_tab not in wb.sheetnames:
            rate_sheet = self.copy_worksheet(wb, self.source_tab, target_tab)
        else:
            rate_sheet = wb[target_tab]
        formatted_rates = self.format_rates(rates)
        self.insert_rates(rate_sheet, formatted_rates, **kwargs)
        self.cleanup_sheet(rate_sheet, rates=rates, formatted_rates=formatted_rates)
        return wb
                
    def insert_rates(self, rate_sheet: xl.worksheet.worksheet.Worksheet,
                     formatted_rates: pd.DataFrame, **kwargs):
        for i, (df_i, row) in zip(range(len(formatted_rates)), formatted_rates.iterrows()):
            for j, val in enumerate(row):
                cell = rate_sheet.cell(i+self.rate_start_row, j+self.rate_start_col)
                cell.value = val
                cell.style = self.rate_style_logic(i, j)
                
    def insert_names(self, rate_sheet: xl.worksheet.worksheet.Worksheet,
                     rates:FillerInputModel, formatted_rates: pd.DataFrame):
        cust_name = rates.cust_name
        for cell in self.name_cells:
            rate_sheet[cell].value = f'{cust_name} - 2023'

    def freeze_start_cell(self, ws: xl.worksheet.worksheet.Worksheet):
        start_col_letter = xl.utils.cell.get_column_letter(self.rate_start_col)
        cell_name = f'{start_col_letter}{self.rate_start_row}'
        ws.freeze_panes = cell_name
                
    def rate_style_logic(self, row_i, col_j):
        return 'RateEven' if (row_i % 2 == 0) else 'RateOdd'
    
    def copy_worksheet(self, source_wb: xl.workbook.workbook.Workbook,
                       source_sheet_name: str, new_sheet_name: str) \
        -> xl.worksheet.worksheet.Worksheet:
        source_sheet = source_wb[source_sheet_name]
        new_sheet = source_wb.copy_worksheet(source_sheet)
        self.copy_all_images(source_sheet, new_sheet)
        new_sheet.title = new_sheet_name
        return new_sheet
        
    def copy_all_images(self, source: xl.worksheet.worksheet.Worksheet,
                        target: xl.worksheet.worksheet.Worksheet):
        new_images = []
        for old_image in source._images:
            new_image = xl.drawing.image.Image(old_image.ref)
            new_image.anchor = old_image.anchor
            new_image.height = old_image.height
            new_image.width = old_image.width
            new_images.append(new_image)
        for image in new_images:
            target.add_image(image)
            
    def cleanup_sheet(self, ws: xl.worksheet.worksheet.Worksheet, rates: FillerInputModel, formatted_rates: pd.DataFrame):
        pass
    
    @abstractmethod
    def format_rates(self, rates: FillerInputModel) -> pd.DataFrame:
        pass
            
            
class WeightBreakFiller(RateTabFiller):
    
    def __init__(self, rate_start_row: int, rate_start_col:int, 
                 source_tab:str, name_width:int, name_row: int,
                 name_space:int):
        super().__init__(rate_start_row=rate_start_row, rate_start_col=rate_start_col,
                         name_cells=['D5', 'T5'], source_tab=source_tab)
        self.name_width = name_width
        self.name_row = name_row
        self.name_space = name_space
        
    def insert_columns(self, rate_sheet: xl.worksheet.worksheet.Worksheet,
                     rates: FillerInputModel, formatted_rates: pd.DataFrame):
        cty_names = rates.zone_map.set_index('ORIGINAL_CTY').COUNTRY_NAME.to_dict()
        for i, col in enumerate(formatted_rates.columns):
            cell = rate_sheet.cell(self.rate_start_row - 1, 
                                   self.rate_start_col + i)
            cell.value = col
            cell.style = 'CountryHeader'
            
            cell_name = rate_sheet.cell(self.rate_start_row - 2, 
                                        self.rate_start_col + i)
            cell_name.value = cty_names.get(col)
            cell_name.style = 'CountryHeader'
    
    def insert_weights(self, rate_sheet: xl.worksheet.worksheet.Worksheet,
                     rates: FillerInputModel, formatted_rates: pd.DataFrame):
        for i, wt in enumerate(formatted_rates.index):
            cell = rate_sheet.cell(self.rate_start_row + i, 
                                   self.rate_start_col - 1)
            cell.value = wt
            cell.style = 'WeightEven' if (i % 2 == 0) else 'WeightOdd'
            
    def create_name_cells(self, rate_sheet: xl.worksheet.worksheet.Worksheet,
                     rates: FillerInputModel, formatted_rates: pd.DataFrame) -> List[str]:        
        num_cells = (formatted_rates.shape[1]//(self.name_space + self.name_width)
                     + (((formatted_rates.shape[1] % (self.name_space + self.name_width))
                         >= self.name_width)))
        num_cells = max(num_cells, 1)
                        
        new_name_cells = []
        for i in range(num_cells):
            start_pos = i * (self.name_width + self.name_space + 1) + self.rate_start_col
            rate_sheet.merge_cells(start_row=self.name_row, 
                                   end_row=self.name_row,
                                   start_column=start_pos,
                                   end_column=start_pos + self.name_width)
            start_col_letter = xl.utils.cell.get_column_letter(start_pos)
            cell_name = f'{start_col_letter}{self.name_row}'
            rate_sheet[cell_name].style = 'WeightOdd'
            new_name_cells.append(cell_name)
        return new_name_cells
    
    def insert_names(self, rate_sheet: xl.worksheet.worksheet.Worksheet,
                     rates:FillerInputModel, formatted_rates: pd.DataFrame):
        cust_name = rates.cust_name
        new_name_cells = self.create_name_cells(rate_sheet, rates, formatted_rates)
        for cell in new_name_cells:
            rate_sheet[cell].value = f'{rates.service_name} - {cust_name}'
            
    def cleanup_sheet(self, ws: xl.worksheet.worksheet.Worksheet, rates: FillerInputModel, formatted_rates: pd.DataFrame):
        self.insert_names(ws, rates=rates, formatted_rates=formatted_rates)
        self.insert_columns(ws, rates, formatted_rates)
        self.insert_weights(ws, rates, formatted_rates)
        self.freeze_start_cell(ws)
        
    def format_rates(self, rates: FillerInputModel) -> pd.DataFrame:
        pivot_rates = rates.base_rates.pivot('PC_WT_MAX', 'ORIGINAL_CTY', 'PC_RATE')
        if (type(rates.zone_map) == pd.DataFrame) and (not rates.zone_map.empty):
            cty_order = (rates.zone_map
                             .sort_values(['ZONE_ORDER', 'ORIGINAL_CTY'])
                             ['ORIGINAL_CTY'].tolist())
            cty_present = [cty for cty in cty_order if cty in pivot_rates.columns]
            return (pivot_rates[cty_present])
        else:
            return pivot_rates
        
        
class PcLbZoneFiller(RateTabFiller):
    def __init__(self, zones_to_int: bool=False, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.zones_to_int = zones_to_int
        
    def format_rates(self, rates: FillerInputModel) -> pd.DataFrame:
        return (rates.base_rates
                    .join(rates.zone_map.set_index('ORIGINAL_CTY'), 
                                   on='ORIGINAL_CTY', rsuffix='_zone')
                    .dropna(subset=['ZONE_CODE'])
                    .sort_values('ZONE_ORDER')
                    [['ZONE_CODE', 'PC_RATE', 'WT_RATE']]
                    .drop_duplicates()
                    .set_index('ZONE_CODE'))
    
    def cleanup_sheet(self, ws: xl.worksheet.worksheet.Worksheet, rates: FillerInputModel, formatted_rates: pd.DataFrame):
        self.insert_names(ws, rates=rates, formatted_rates=formatted_rates)
        self.insert_zones(ws, rates, formatted_rates)
        
    def insert_zones(self, ws: xl.worksheet.worksheet.Worksheet,
                     rates:FillerInputModel, formatted_rates: pd.DataFrame):
        zone_names = rates.zone_map.set_index('ZONE_CODE').ZONE_DESCRIPTION.to_dict()
        for i, zone in enumerate(formatted_rates.index):
            curr_style = 'ZoneEven' if (i % 2) == 0 else 'ZoneOdd'
            cell = ws.cell(self.rate_start_row + i, 
                                   self.rate_start_col - 1)
            cell.value = zone_names.get(zone)
            cell.style = curr_style
            
            cell_name = ws.cell(self.rate_start_row + i, 
                                self.rate_start_col - 2)
            cell_name.value = self.try_convert_int(zone) if self.zones_to_int else zone
            cell_name.style = curr_style
    
    def try_convert_int(self, val):
        try:
            return int(val)
        except:
            return val
        
        
class WeightBreakZoneFiller(RateTabFiller):
    def __init__(self, rate_start_row: int, rate_start_col:int, 
                 source_tab:str, name_row: int, *args, **kwargs):
        super().__init__(rate_start_row=rate_start_row, rate_start_col=rate_start_col,
                         name_cells=['D5', 'T5'], source_tab=source_tab)
        self.name_row = name_row
        
    def format_rates(self, rates: FillerInputModel) -> pd.DataFrame:
        zone_order = (rates.zone_map
                        .sort_values(['ZONE_ORDER', 'ZONE_CODE'])
                        ['ZONE_CODE'].drop_duplicates().tolist())
                        
        pivot_rates = (rates.base_rates
                    .join(rates.zone_map.set_index('ORIGINAL_CTY'), 
                           on='ORIGINAL_CTY', rsuffix='_zone')
                    .dropna(subset=['ZONE_CODE'])
                    [['ZONE_CODE', 'PC_RATE', 'PC_WT_MAX']]
                    .drop_duplicates()
                    .pivot(index='PC_WT_MAX', columns='ZONE_CODE', values='PC_RATE'))

        zones_present = [zone for zone in zone_order if zone in pivot_rates.columns]
        return pivot_rates[zones_present]
    
    def cleanup_sheet(self, ws: xl.worksheet.worksheet.Worksheet, rates: FillerInputModel, formatted_rates: pd.DataFrame):
        self.insert_names(ws, rates=rates, formatted_rates=formatted_rates)
        self.insert_zones(ws, rates, formatted_rates)
        self.insert_weights(ws, rates, formatted_rates)
        
    def insert_names(self, ws, rates, formatted_rates):
        name_length = formatted_rates.shape[1]
        
        #always one extra column left for service name
        ws.merge_cells(start_row=self.name_row, 
                        end_row=self.name_row,
                        start_column=self.rate_start_col - 1,
                        end_column=self.rate_start_col + name_length - 1)
        #one above is zone label, two above is zone description, 3 above
        ws.merge_cells(start_row=self.rate_start_row - 3, 
                        end_row=self.rate_start_row - 3,
                        start_column=self.rate_start_col - 1,
                        end_column=self.rate_start_col + name_length - 1)

        start_col_letter = xl.utils.cell.get_column_letter(self.rate_start_col - 1)
        cell_name = f'{start_col_letter}{self.name_row}'
        cell_service = f'{start_col_letter}{self.rate_start_row - 3}'
        ws[cell_name].style = 'NameHeader'
        ws[cell_name].value = rates.cust_name
        ws[cell_service].style = 'ServiceHeader'
        ws[cell_service].value = rates.service_name
        
        
    def insert_zones(self, rate_sheet: xl.worksheet.worksheet.Worksheet,
                     rates: FillerInputModel, formatted_rates: pd.DataFrame):
        zone_names = (rates.zone_map[['ZONE_CODE', 'ZONE_DESCRIPTION']]
                          .drop_duplicates()
                          .set_index('ZONE_CODE').ZONE_DESCRIPTION.to_dict())
        for i, zone in enumerate(formatted_rates.columns):
            cell = rate_sheet.cell(self.rate_start_row - 1, 
                                   self.rate_start_col + i)
            cell.value = zone
            cell.style = 'CountryHeader'
            
            cell_name = rate_sheet.cell(self.rate_start_row - 2, 
                                        self.rate_start_col + i)
            cell_name.value = zone_names.get(zone)
            cell_name.style = 'CountryHeader'
    
    def insert_weights(self, rate_sheet: xl.worksheet.worksheet.Worksheet,
                     rates: FillerInputModel, formatted_rates: pd.DataFrame):
        for i, wt in enumerate(formatted_rates.index):
            cell = rate_sheet.cell(self.rate_start_row + i, 
                                   self.rate_start_col - 1)
            cell.value = wt
            cell.style = 'WeightEven' if (i % 2 == 0) else 'WeightOdd'
    
    def try_convert_int(self, val):
        try:
            return int(val)
        except:
            return val