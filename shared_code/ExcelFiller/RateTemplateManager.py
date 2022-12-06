# -*- coding: utf-8 -*-
"""
Created on Thu Jul 21 14:19:26 2022

@author: rtse
"""

from shared_code.ExcelFiller.RateTabFiller import RateTabFiller, WeightBreakFiller, PcLbZoneFiller, WeightBreakZoneFiller
from shared_code.ExcelFiller.FillerInput import FillerInputModel, QuoteParamsModel
import openpyxl as xl
import pandas as pd
from io import BytesIO
from dataclasses import dataclass
from typing import List, Dict, Tuple
from abc import abstractmethod, ABC


@dataclass
class FillerLogicUnit:
    target_tab: str
    tab_filler: RateTabFiller
    aux_start_tabs: List[str] = None
    aux_end_tabs: List[str] = None


class RateCardManager(ABC):
    
    def __init__(self, template_path: str, template_source_tab: str = None,
                 start_tabs: List[str] = None, end_tabs: List[str] = None, 
                 aux_start_tabs_map: Dict[int, List[str]]=None, aux_end_tabs_map:Dict[int, List[str]]=None, *args, **kwargs):
        self.template_path = template_path
        self.template_source_tab = template_source_tab
        with open(self.template_path, 'rb') as f:
            self.template = BytesIO(f.read())
        self.start_tabs = start_tabs if start_tabs else []
        self.end_tabs = end_tabs if end_tabs else []
        self.aux_start_tabs_map = aux_start_tabs_map if aux_start_tabs_map else {}
        self.aux_end_tabs_map = aux_end_tabs_map if aux_end_tabs_map else {}
            
    def save_new_workbook(self, filename: str, filler_inputs: List[FillerInputModel], 
                          quote_params: QuoteParamsModel, in_mem: bool = True, *args, **kwargs):
        wb = xl.load_workbook(self.template)
        start_tabs_used = self.start_tabs.copy()
        fill_tabs_used = []
        end_tabs_used = self.end_tabs.copy()
        for filler_input in filler_inputs:
            filler_logic_unit= self.assign_filler_logic(filler_input)
            filler_logic_unit.tab_filler.fill_tab(wb, filler_input, target_tab=filler_logic_unit.target_tab)
            fill_tabs_used.append(filler_logic_unit.target_tab)
            if filler_logic_unit.aux_end_tabs:
                for tab in filler_logic_unit.aux_end_tabs:
                    #intent is to keep order and uniqueness probably a better way
                    if tab not in end_tabs_used:
                        end_tabs_used.append(tab)
        self.cleanup_workbook(wb, fill_tabs_used, filler_inputs, quote_params, start_tabs_used, end_tabs_used)
        if in_mem:
            return wb
        wb.save(filename)
        
    def cleanup_workbook(self, wb, fill_tabs_used, filler_inputs, quote_params: QuoteParamsModel, start_tabs_used, end_tabs_used, *args, **kwargs):
        ordered_tabs = start_tabs_used + fill_tabs_used + end_tabs_used
        self.delete_unused_tabs(wb, ordered_tabs)
        self.reorder_tabs(wb, ordered_tabs)
        
    def delete_unused_tabs(self, wb: xl.workbook.workbook.Workbook, tabs_used: List[str]):
        for sheet in wb.sheetnames:
            if sheet not in tabs_used:
                del wb[sheet]
                
    def reorder_tabs(self, wb, tabs_used):
        current_index = [wb.sheetnames.index(tab) for tab in tabs_used]
        wb._sheets = [wb._sheets[i] for i in current_index]
        wb._active_sheet_index = 0
                
    @abstractmethod
    def assign_filler_logic(self, filler_input: FillerInputModel) -> FillerLogicUnit:
        pass
    
    
class ShoppedTemplateManager(RateCardManager):
    
    cover_name_cell = 'B2'
    cover_quotenum_cell = 'B3'
    cover_date_cell = 'B4'
    cover_source_tab = 'Service Overview'
    surcharges_source_tab = 'Surcharges'
    surcharges_start_row = 10
    surcharges_start_col = 1
    generic_shopped_sheet = WeightBreakFiller(rate_start_row=9, rate_start_col=4, 
                                              source_tab='Shopped Lb', name_width=15, name_row=5,
                                              name_space=1)
    
    def assign_filler_logic(self, filler_input):
        start_tabs = self.aux_start_tabs_map.get(filler_input.service_id) if self.aux_start_tabs_map else None
        end_tabs = self.aux_end_tabs_map.get(filler_input.service_id) if self.aux_end_tabs_map else None
        return FillerLogicUnit(filler_input.service_abbr, self.generic_shopped_sheet, start_tabs, end_tabs)
    
    def cleanup_workbook(self, wb, fill_tabs_used, filler_inputs, quote_params: QuoteParamsModel, start_tabs_used, end_tabs_used, *args, **kwargs):
        self.fill_cover_page(wb, quote_params, fill_tabs_used)
        default_service_order = ['PPND', 'PPNDP', 'PPNDU', 'PPT', 'PPDC', 'PPDCP', 'PPDCU']
        reordered_services = [svc for svc in default_service_order if svc in fill_tabs_used]
        remaining_backup = sorted([tab for tab in fill_tabs_used if tab not in reordered_services])
        all_reordered_tabs = reordered_services + remaining_backup + [self.cover_source_tab]
        #ugh reconstruct surcharges here and do something similar they're in the filler inputs
        surcharges = filler_inputs[0].surcharges #all fillers now have a copy of all surcharges maybe clean this up later
        if not surcharges.empty:
            self.fill_surcharges_page(wb, surcharges)
            end_tabs_used.append(self.surcharges_source_tab)
        super().cleanup_workbook(wb, all_reordered_tabs, filler_inputs, quote_params, start_tabs_used, end_tabs_used)
        
    def fill_cover_page(self, wb, quote_params, fill_tabs_used):
        cover_sheet = wb[self.cover_source_tab]
        cover_sheet[self.cover_name_cell].value = quote_params.cust_name
        cover_sheet[self.cover_quotenum_cell].value = quote_params.quote_num
        cover_sheet[self.cover_date_cell].value = quote_params.quote_date

        #this definitely shouldn't be here
        services_info_order = [['PPND', 'PPNDP', 'PPNDU'], ['PPT'], ['PPDCP', 'PPDC'], ['PPDCU', 'PPDC']]
        services_info_current_col = 5
        for services_per_col in services_info_order:
            if not any([svc in fill_tabs_used for svc in services_per_col]):
                cover_sheet.delete_cols(services_info_current_col)
            else:
                services_info_current_col += 1
        cover_sheet.move_range("D6:H20", rows=0, cols=-3, translate=True)

    def fill_surcharges_page(self, wb, surcharges):
        surcharge_sheet = wb[self.surcharges_source_tab]
        for i, (df_i, row) in zip(range(len(surcharges)), surcharges.iterrows()):
            for j, val in enumerate(row):
                cell = surcharge_sheet.cell(i+self.surcharges_start_row, j+self.surcharges_start_col)
                cell.value = val
                cell.style = 'RateEven' if (i % 2) == 0 else 'RateOdd'
        
        
class PcLbZoneManager(RateCardManager):
    
    def assign_filler_logic(self, filler_input: FillerInputModel) -> Tuple[str, RateTabFiller]:
        return FillerLogicUnit(self.template_source_tab, PcLbZoneFiller(rate_start_row=11, rate_start_col=4,
                                    name_cells = ['B7'], source_tab= self.template_source_tab,
                                    zones_to_int=True))


class PcLbEPSManager(PcLbZoneManager):
    
    def assign_filler_logic(self, filler_input: FillerInputModel) -> Tuple[str, RateTabFiller]:
        return FillerLogicUnit(self.template_source_tab, PcLbZoneFiller(rate_start_row=11, rate_start_col=4,
                                    name_cells = ['B7'], source_tab= self.template_source_tab,
                                    zones_to_int=True))

class PcLbIPAManager(PcLbZoneManager):
    
    def assign_filler_logic(self, filler_input: FillerInputModel) -> Tuple[str, RateTabFiller]:
        return FillerLogicUnit(self.template_source_tab, PcLbZoneFiller(rate_start_row=12, rate_start_col=4,
                                    name_cells = ['B7'], source_tab= self.template_source_tab,
                                    zones_to_int=True))
    
    
class WtBreakZoneManager(RateCardManager):
    
    def assign_filler_logic(self, filler_input: FillerInputModel) -> Tuple[str, RateTabFiller]:
        return FillerLogicUnit(self.template_source_tab, WeightBreakZoneFiller(rate_start_row=13, rate_start_col=2,
                                              name_cells = ['A8'], source_tab=self.template_source_tab,
                                              name_row=8))